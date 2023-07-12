import pytest
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import time
import openpyxl

@pytest.fixture()
def driver():
    driver = webdriver.Chrome()
    driver.get('https://opensource-demo.orangehrmlive.com/')
    driver.maximize_window()
    yield driver
    driver.quit()


def test(driver):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    
    # Login
    time.sleep(3)
    driver.find_element(By.NAME, 'username').send_keys('Admin')
    driver.find_element(By.NAME, 'password').send_keys('admin123' + Keys.ENTER)
    time.sleep(3)
    
    worksheet.cell(row=1, column=1).value = 'Login Step'
    worksheet.cell(row=1, column=2).value = 'Pass'

    # Masuk Menu Admin
    driver.find_element(By.XPATH, '//*[@id="app"]/div[1]/div[1]/aside/nav/div[2]/ul/li[1]/a').click()
    time.sleep(3)
    
    worksheet.cell(row=2, column=1).value = 'Masuk Menu Admin'
    worksheet.cell(row=2, column=2).value = 'Pass'

    # Mencari User menggunakan username Benar
    driver.find_element(By.XPATH, '//*[@id="app"]/div[1]/div[2]/div[2]/div/div[1]/div[2]/form/div[1]/div/div[1]/div/div[2]/input').send_keys('Jasmine.Morgan')
    driver.find_element(By.XPATH, '//*[@id="app"]/div[1]/div[2]/div[2]/div/div[1]/div[2]/form/div[2]/button[2]').click()
    time.sleep(3)
    driver.find_element(By.XPATH, '//*[@id="app"]/div[1]/div[2]/div[2]/div/div[1]/div[2]/form/div[2]/button[1]').click()

    # Mencari User menggunakan username Salah
    driver.find_element(By.XPATH, '//*[@id="app"]/div[1]/div[2]/div[2]/div/div[1]/div[2]/form/div[1]/div/div[1]/div/div[2]/input').send_keys('Raiyana')
    driver.find_element(By.XPATH, '//*[@id="app"]/div[1]/div[2]/div[2]/div/div[1]/div[2]/form/div[2]/button[2]').click()
    time.sleep(3)
    driver.find_element(By.XPATH, '//*[@id="app"]/div[1]/div[2]/div[2]/div/div[1]/div[2]/form/div[2]/button[1]').click()

    # Menampilkan data User tanpa memasukkan input apapun
    driver.find_element(By.XPATH, '//*[@id="app"]/div[1]/div[2]/div[2]/div/div[1]/div[2]/form/div[2]/button[2]').click()
    time.sleep(3)
    driver.find_element(By.XPATH, '//*[@id="app"]/div[1]/div[2]/div[2]/div/div[1]/div[2]/form/div[2]/button[1]').click()

    # Menambah User baru tanpa mengisi input apapun
    driver.find_element(By.XPATH, '//*[@id="app"]/div[1]/div[2]/div[2]/div/div[2]/div[1]/button').click()
    time.sleep(3)
    driver.find_element(By.XPATH, '//*[@id="app"]/div[1]/div[2]/div[2]/div/div/form/div[3]/button[2]').click()
    time.sleep(3)
    driver.find_element(By.XPATH, '//*[@id="app"]/div[1]/div[2]/div[2]/div/div/form/div[3]/button[1]').click()
    time.sleep(3)

    # Add more steps and their corresponding results to the worksheet as needed

    workbook.save('admin_output2.xlsx')
